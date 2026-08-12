#!/usr/bin/env python3
"""Genera ranking.json a partir del Excel local de jugadores.

Uso:
    python3 update_ranking.py [ruta_al_excel]

Si no se indica ruta, usa EXCEL_PATH.
"""
import json
import sys
from pathlib import Path

from openpyxl import load_workbook

EXCEL_PATH = Path("/Users/jiajia.ye/Desktop/Garpadel-ranking/jugadores.xlsx")
OUTPUT_PATH = Path(__file__).parent / "ranking.json"


def main() -> None:
    excel_path = Path(sys.argv[1]) if len(sys.argv) > 1 else EXCEL_PATH
    if not excel_path.exists():
        sys.exit(f"No se encuentra el Excel: {excel_path}")

    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active

    headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
    try:
        col = {name: headers.index(name) for name in ("nombre", "nivel", "puntuacion")}
    except ValueError as e:
        sys.exit(f"Falta una columna esperada (nombre, nivel, puntuacion) en {excel_path}: {e}")

    jugadores = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        nombre = row[col["nombre"]]
        if not nombre:
            continue
        jugadores.append(
            {
                "nombre": str(nombre),
                "nivel": str(row[col["nivel"]] or ""),
                "puntuacion": int(row[col["puntuacion"]] or 0),
            }
        )

    jugadores.sort(key=lambda j: j["puntuacion"], reverse=True)

    OUTPUT_PATH.write_text(
        json.dumps(jugadores, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(f"OK: {OUTPUT_PATH} actualizado con {len(jugadores)} jugadores")


if __name__ == "__main__":
    main()
